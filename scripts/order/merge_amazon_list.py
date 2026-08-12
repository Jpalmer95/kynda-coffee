#!/usr/bin/env python3
"""
Merge exact Amazon listing titles (from amazon_lists_raw.json) with the par
levels from the MONTHLY par sheet, producing the authoritative
amazon_canonical_list.tsv keyed to Amazon's exact product names + ASINs.

Matching is fuzzy (par-sheet names are hand-typed approximations). Unmatched
items are flagged for manual review.

Output: name | par | Amazon | category(food|packaging) | asin
"""
import json, re, sys, unicodedata

def norm(s):
    s = unicodedata.normalize("NFKD", s).lower()
    s = s.replace("®", "").replace("™", "").replace("©", "")
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    s = re.sub(r"\b(oz|lb|lbs|ct|pk|pack|packs?|case|count|fl|x|with|for|of|the|and|by|amp)\b", " ", s)
    s = re.sub(r"\d+(?:\.\d+)?", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def tokens(s):
    return set(norm(s).split())

def score(a, b):
    A, B = tokens(a), tokens(b)
    if not A or not B: return 0
    hits = sum(1 for w in A if w in B)
    return hits / max(len(A), 1)

def load_par_sheet(path="/mnt/flex/Kynda 2026 Drive Download 8.10.2026/KYNDA COFFEE 2026/05_Inventory/par-and-waste/Inventory Par Sheet 2026_.xlsx"):
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb["MONTHLY Inventory"]
    items = []
    for r in range(10, 300):
        a = ws.cell(r, 1).value
        if a is None or str(a).strip() == "":
            continue
        av = str(a).strip()
        if av.isupper() and any(k in av.upper() for k in ["HEB", "NUMINOUS", "BROTHER", "PRODUCT"]):
            break
        if "AMAZON" in av: continue
        stock, par = ws.cell(r, 3).value, ws.cell(r, 4).value
        if stock is None and par is None: continue
        items.append({"name": av, "par": par if par is not None else ""})
    return items

def main():
    raw = json.load(open("/tmp/amazon_lists_raw.json"))
    pars = load_par_sheet()
    print(f"Par-sheet items: {len(pars)} | Amazon food: {len(raw['food'])} | packaging: {len(raw['packaging'])}")

    out = []
    unmatched_amazon = []
    used_par = set()

    for cat, arr in [("food", raw["food"]), ("packaging", raw["packaging"])]:
        for it in arr:
            t = it["title"].strip()
            if not t or len(t) < 8 or t.startswith(("-", "$", "%")):
                continue
            # Amazon titles can contain '|' — replace so it doesn't break the TSV
            t = t.replace("|", " - ").strip()
            # find best par-sheet match
            best, bs = None, 0
            for i, p in enumerate(pars):
                if i in used_par: continue
                s = score(t, p["name"])
                if s > bs: bs, best = s, p
            par_val = ""
            if best and bs >= 0.35:
                par_val = best["par"]
                used_par.add(pars.index(best))
            out.append(f"{t}|{par_val}|Amazon|{cat}|{it['asin']}")

    # Report par-sheet items not matched
    unmatched_par = [p["name"] for i, p in enumerate(pars) if i not in used_par]

    with open("scripts/order/amazon_canonical_list.tsv", "w") as f:
        f.write("# Kynda AMAZON — CANONICAL ordering list (exact Amazon listing titles + ASINs)\n")
        f.write("# Auto-merged from live Amazon 'Kynda Food List' + 'Kynda Packaging + Supplies List'\n")
        f.write("# Format: name | par | vendor | category | asin\n")
        for line in out:
            f.write(line + "\n")

    print(f"Wrote {len(out)} items to amazon_canonical_list.tsv")
    print(f"\nUnmatched par-sheet items ({len(unmatched_par)}) — verify names:")
    for p in unmatched_par:
        print(f"  • {p}")

if __name__ == "__main__":
    main()
